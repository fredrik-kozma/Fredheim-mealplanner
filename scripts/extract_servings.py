"""Scan the Standard Recipes 2022 spreadsheets and collect servings per title."""
import openpyxl, os, re, json, sys

SS_DIR = r'C:\Users\fredr\OneDrive\Skrivbord\Standard recipes 2022 (spreadsheets)'
PACK = 'recipe-packs-template/packs/fredheim-recipes.json'

def extract(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return None, None, str(e)
    # INSTITUTION has the canonical title + servings. ENGLISH is sometimes
    # blank or a rename. Read INSTITUTION first; fall back to ENGLISH.
    preferred = ['Recipe INSTITUTION', 'Recipe ENGLISH']
    sheet_order = [s for s in preferred if s in wb.sheetnames] + \
                  [s for s in wb.sheetnames if s not in preferred]
    title = None
    servings = None
    for s in sheet_order:
        ws = wb[s]
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
            if not row: continue
            key = (row[0] or '').strip().upper() if row[0] else ''
            val = row[1] if len(row) > 1 else None
            if key.startswith('TITLE') and val and title is None:
                title = str(val).strip()
            if key.startswith('SERVINGS') and val is not None and servings is None:
                try: servings = int(float(val))
                except Exception: pass
        if title and servings: break
    wb.close()
    return title, servings, None

def norm(s):
    s = (s or '').lower()
    s = re.sub(r'[^a-z0-9]+', ' ', s).strip()
    return s

def main():
    rows = []
    files = [fn for fn in sorted(os.listdir(SS_DIR)) if fn.lower().endswith('.xlsx')]
    for i, fn in enumerate(files):
        full = os.path.join(SS_DIR, fn)
        title, servings, err = extract(full)
        rows.append((fn, title, servings, err))
        if (i+1) % 25 == 0:
            print(f"  processed {i+1}/{len(files)}", flush=True)

    missing_title = [r for r in rows if not r[1]]
    missing_serv  = [r for r in rows if r[1] and r[2] is None]
    print(f"\nTotal files: {len(rows)}")
    print(f"Missing title: {len(missing_title)}")
    print(f"Title but no servings: {len(missing_serv)}")
    for r in missing_serv:
        print(" no servings:", r[1], '—', r[0])

    lookup = {}
    for fn, title, servings, err in rows:
        if title and servings:
            lookup[norm(title)] = (title, servings, fn)

    with open(PACK, encoding='utf-8') as f:
        pack = json.load(f)

    matched, unmatched = 0, []
    for r in pack['recipes']:
        k = norm(r['title'])
        if k in lookup: matched += 1
        else: unmatched.append(r['title'])
    print(f"\nPack recipes: {len(pack['recipes'])}")
    print(f"Matched: {matched}")
    print(f"Unmatched ({len(unmatched)}):")
    for t in unmatched: print("  ", t)

    # Save intermediate lookup for re-use
    with open('scripts/_tmp/servings_lookup.json', 'w', encoding='utf-8') as f:
        json.dump(
            {k: {'title': v[0], 'servings': v[1], 'file': v[2]} for k, v in lookup.items()},
            f, ensure_ascii=False, indent=2,
        )
    print("\nSaved scripts/_tmp/servings_lookup.json")

if __name__ == '__main__':
    os.makedirs('scripts/_tmp', exist_ok=True)
    main()
